import os
import sys
import ast
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils.cell import get_column_letter
import google.generativeai as genai


from dotenv import load_dotenv
load_dotenv()
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))


def input_image_setup(file_location, type='image/jpeg'):
    # Check if the file location is valid
    if file_location:
            # Open the file in binary mode and read it into bytes
        with open(file_location, 'rb') as f:
                bytes_data = f.read()
        print("Data Extraction Started..")
        image_parts = [
                {
                    "mime_type": type,  # Get the mime type of the uploaded file
                    "data": bytes_data
                }
            ]

        input_prompt="""
                You are an expert in reading handwriting in english. There is a table present in the provided image. Read the image and extract the table data which include Names , phone number and email . Try your best to get the best possible result. The handwriting will in different in every rows. Identify the bounding lines to extract information row wise.
                The table should be in the following format:
                Format should not be changed anyhow.
                [[Column1,Column2,...],[Data1,Data2,Data3...],[]...]
            """

        generation_config = {
            "temperature": 0.1,
            "top_p": 1,
            "top_k": 1,
            "max_output_tokens": 2048,
            }
        model=genai.GenerativeModel('gemini-pro-vision',generation_config=generation_config)
        response=model.generate_content([input_prompt,image_parts[0]])
        # print(response.text)
        print("Data Extracted Successfully!")
        return response.text



def process(list_a):
    print("Data Processing...")
    Data=[]
    columns=[]
    columns.append(list_a[0][1:4])
    for i in range(1,len(list_a)):
        Data.append(list_a[i][1:4])
    d = np.array(Data)
    return d,columns



def data_to_xlsx(data, columns, filename):
    # Create a pandas DataFrame from the data
    print("Creating Excel...")
    df = pd.DataFrame(data, columns=columns)
    # print(df)
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers to the first row
    for col_num, header in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data to subsequent rows
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to a file
    workbook.save(filename)

if __name__ == "__main__":
    if len(sys.argv)!= 2:
        print("Usage: python script.py <file_location>")
        sys.exit(1)
    file_location = sys.argv[1]
    Raw=input_image_setup(file_location)
    Raw = Raw.replace("\t", "").replace(" ", "")
    list_a = ast.literal_eval(Raw)
    d,c = process(list_a)
    data_to_xlsx(d, c[0], "output/output.xlsx")
    print("Script Success !")